from openpyxl import load_workbook
from ru.textilstock.assortment.mixbox.MixBox import MixBox, MixBoxItem


def is_mix_box_start(row):
    val = row[0].value
    return val is not None and isinstance(val, str) and val.startswith('Поставщик: ')


def is_mix_box_end(row):
    val = row[2].value
    return val is not None and isinstance(val, str) and val == 'ИТОГО В КОРОБЕ'


def prepare_row_boxes(ws):
    row_boxes = []
    cur_rowboxes = None
    for row in ws.iter_rows():
        if is_mix_box_start(row):
            cur_rowboxes = []

        if is_mix_box_end(row):
            cur_rowboxes.append(row)
            row_boxes.append(cur_rowboxes)
            cur_rowboxes = None

        if cur_rowboxes is not None:
            cur_rowboxes.append(row)
    return row_boxes


def process_box(row_box):
    first_item_row_offset = 6
    supplier_name = row_box[0][0].value[11:]
    brand = row_box[1][0].value[7:]
    consignee = row_box[2][0].value[17:]
    cargo_number_str = row_box[3][0].value[12:]
    try:
        cargo_number = int(cargo_number_str)
    except ValueError:
        print('BAD CARGO NUM. SKIPPED')
        return None
    num_and_count = row_box[4][2].value.split()
    box_num = int(num_and_count[0])
    box_count = int(num_and_count[2])
    sum_items_count = int(row_box[len(row_box) - 1][4].value)

    item_rows = row_box[first_item_row_offset:len(row_box) - 1]

    items = []
    for item_row in item_rows:
        item = MixBoxItem(item_row[0].value,
                          item_row[1].value,
                          item_row[2].value,
                          item_row[3].value,
                          item_row[4].value)
        items.append(item)

    box = MixBox(supplier_name, brand, consignee, cargo_number,
                 box_num, box_count, items, sum_items_count)
    print('Proeessed. cargo num: ' + str(box.cargo_number)
          + ', box num: ' + str(box_num))
    return box


"""
Returns MixBox array
 :rtype: :class:`MixBox`
"""


def parse(mix_xlsx_file_name):
    workbook = load_workbook(filename=mix_xlsx_file_name, data_only=True)
    ws = workbook.active
    row_boxes = prepare_row_boxes(ws)
    result = []
    for rb in row_boxes:
        box = process_box(rb)
        if box is not None:
            result.append(box)
    return result

