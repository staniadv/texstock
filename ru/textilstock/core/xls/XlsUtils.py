import xlrd
from openpyxl.workbook import Workbook

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(1, nrows):
        for col in range(1, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return book1

def print_all_rows(ws):
    for row in ws.iter_rows():
        for cell in row:
            print(cell.value, end=' ')
        print()
